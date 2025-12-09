"""
Excel file generator for testing.
Generates realistic Excel files without manual file creation.
"""
from datetime import date, timedelta
from io import BytesIO
from typing import Optional, List, Dict, Any
from openpyxl import Workbook


class ExcelGenerator:
    """Generate realistic Excel files for testing import functionality."""
    
    @staticmethod
    def create_simple_import(
        num_tasks: int = 10,
        program_id: str = "PROG-001",
        project_id: str = "PROJ-001",
        phase_id: str = "PHASE-001"
    ) -> bytes:
        """
        Create Excel with N simple tasks.
        
        Args:
            num_tasks: Number of tasks to generate
            program_id: Program external ID
            project_id: Project external ID
            phase_id: Phase external ID
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Items"
        
        # Headers matching expected format
        headers = [
            "Program ID", "Program Name", "Project ID", "Project Name",
            "Phase ID", "Phase Name", "Phase Sequence", "Work Item ID",
            "Work Item Name", "Planned Start", "Planned End", "Planned Effort",
            "Assigned Resource", "Complexity Level", "Revenue Impact $",
            "Strategic Importance", "Customer Impact", "Critical for Launch?",
            "Feature Name"
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Generate tasks
        base_date = date.today()
        for i in range(num_tasks):
            row = i + 2
            task_start = base_date + timedelta(days=i * 5)
            task_end = task_start + timedelta(days=10)
            
            values = [
                program_id,                           # Program ID
                "Test Program",                       # Program Name
                project_id,                           # Project ID
                "Test Project",                       # Project Name
                phase_id,                             # Phase ID
                "Phase 1",                            # Phase Name
                1,                                    # Phase Sequence
                f"TASK-{i+1:04d}",                   # Work Item ID
                f"Test Task {i+1}",                   # Work Item Name
                task_start.isoformat(),               # Planned Start
                task_end.isoformat(),                 # Planned End
                40,                                   # Planned Effort
                f"RES-{(i % 5) + 1:03d}" if i % 3 == 0 else None,  # Resource
                ["Low", "Medium", "High"][i % 3],     # Complexity
                (i + 1) * 1000 if i % 2 == 0 else None,  # Revenue Impact
                ["High", "Medium", "Low"][i % 3],     # Strategic Importance
                ["High", "Medium", "Low"][i % 3],     # Customer Impact
                i % 4 == 0,                           # Critical for Launch
                f"Feature-{(i % 3) + 1}" if i % 2 == 0 else None,  # Feature Name
            ]
            
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
        
        # Add Resources sheet
        ws_resources = wb.create_sheet("Resources")
        resource_headers = [
            "Resource ID", "Resource Name", "Email", "Role",
            "Home Program/Team", "Cost Per Hour", "Max Utilization",
            "Skill Level", "Location"
        ]
        for col, header in enumerate(resource_headers, start=1):
            ws_resources.cell(row=1, column=col, value=header)
        
        # Generate 5 resources
        for i in range(5):
            row = i + 2
            values = [
                f"RES-{i+1:03d}",
                f"Test User {i+1}",
                f"user{i+1}@example.com",
                ["Developer", "Designer", "PM", "QA", "DevOps"][i],
                "Test Team",
                100 + i * 10,
                100,
                ["Junior", "Mid", "Senior", "Lead", "Principal"][i],
                "Remote"
            ]
            for col, value in enumerate(values, start=1):
                ws_resources.cell(row=row, column=col, value=value)
        
        # Add Dependencies sheet (empty initially)
        ws_deps = wb.create_sheet("Dependencies")
        dep_headers = [
            "Successor Task ID", "Predecessor Task ID",
            "Dependency Type", "Lag Days", "Notes"
        ]
        for col, header in enumerate(dep_headers, start=1):
            ws_deps.cell(row=1, column=col, value=header)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def create_large_import(
        num_tasks: int = 5000,
        program_id: str = "PROG-001"
    ) -> bytes:
        """
        Create Excel with many tasks for performance testing.
        
        Args:
            num_tasks: Number of tasks (default 5000)
            program_id: Program external ID
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Items"
        
        # Headers
        headers = [
            "Program ID", "Program Name", "Project ID", "Project Name",
            "Phase ID", "Phase Name", "Phase Sequence", "Work Item ID",
            "Work Item Name", "Planned Start", "Planned End", "Planned Effort"
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Generate tasks across multiple projects and phases
        base_date = date.today()
        projects_per_program = 10
        phases_per_project = 5
        
        for i in range(num_tasks):
            row = i + 2
            project_num = (i // 100) % projects_per_program + 1
            phase_num = (i // 20) % phases_per_project + 1
            task_start = base_date + timedelta(days=(i % 100))
            task_end = task_start + timedelta(days=10)
            
            values = [
                program_id,
                "Large Test Program",
                f"PROJ-{project_num:03d}",
                f"Project {project_num}",
                f"PHASE-{project_num:03d}-{phase_num:02d}",
                f"Phase {phase_num}",
                phase_num,
                f"TASK-{i+1:06d}",
                f"Task {i+1}: Implementation Item",
                task_start.isoformat(),
                task_end.isoformat(),
                40
            ]
            
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def create_with_dependencies(
        tasks: List[Dict[str, Any]],
        dependencies: List[Dict[str, str]],
        program_id: str = "PROG-001"
    ) -> bytes:
        """
        Create Excel with specific tasks and dependencies.
        
        Args:
            tasks: List of task dictionaries
            dependencies: List of dependency dictionaries
            program_id: Program external ID
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Items"
        
        headers = [
            "Program ID", "Program Name", "Project ID", "Project Name",
            "Phase ID", "Phase Name", "Phase Sequence", "Work Item ID",
            "Work Item Name", "Planned Start", "Planned End", "Planned Effort"
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        for i, task in enumerate(tasks):
            row = i + 2
            values = [
                program_id,
                "Test Program",
                task.get("project_id", "PROJ-001"),
                task.get("project_name", "Test Project"),
                task.get("phase_id", "PHASE-001"),
                task.get("phase_name", "Phase 1"),
                task.get("phase_sequence", 1),
                task["external_id"],
                task["name"],
                task.get("planned_start", date.today().isoformat()),
                task.get("planned_end", (date.today() + timedelta(days=10)).isoformat()),
                task.get("effort", 40)
            ]
            
            for col, value in enumerate(values, start=1):
                ws.cell(row=row, column=col, value=value)
        
        # Add Dependencies sheet
        ws_deps = wb.create_sheet("Dependencies")
        dep_headers = [
            "Successor Task ID", "Predecessor Task ID",
            "Dependency Type", "Lag Days", "Notes"
        ]
        for col, header in enumerate(dep_headers, start=1):
            ws_deps.cell(row=1, column=col, value=header)
        
        for i, dep in enumerate(dependencies):
            row = i + 2
            values = [
                dep["successor_id"],
                dep["predecessor_id"],
                dep.get("type", "FS"),
                dep.get("lag_days", 0),
                dep.get("notes", "")
            ]
            for col, value in enumerate(values, start=1):
                ws_deps.cell(row=row, column=col, value=value)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def create_smart_merge_test_data(
        existing_task_ids: List[str],
        new_task_ids: List[str],
        removed_task_ids: List[str],  # Tasks that exist in DB but NOT in this Excel
        program_id: str = "PROG-001"
    ) -> bytes:
        """
        Create Excel for testing Smart Merge scenarios.
        
        Case A (Insert): new_task_ids - tasks in Excel but not in DB
        Case B (Update): existing_task_ids - tasks in both Excel and DB
        Case C (Ghost): removed_task_ids - tasks in DB but not in Excel
        
        Args:
            existing_task_ids: Task IDs that exist in DB (for update)
            new_task_ids: New task IDs (for insert)
            removed_task_ids: Not included in Excel (for ghost check)
            program_id: Program external ID
            
        Returns:
            Excel file as bytes
        """
        # Combine existing and new (exclude removed)
        all_tasks = []
        base_date = date.today()
        
        for i, task_id in enumerate(existing_task_ids):
            all_tasks.append({
                "external_id": task_id,
                "name": f"Existing Task: {task_id}",
                "planned_start": (base_date + timedelta(days=i * 5)).isoformat(),
                "planned_end": (base_date + timedelta(days=i * 5 + 10)).isoformat(),
            })
        
        for i, task_id in enumerate(new_task_ids):
            all_tasks.append({
                "external_id": task_id,
                "name": f"New Task: {task_id}",
                "planned_start": (base_date + timedelta(days=i * 3)).isoformat(),
                "planned_end": (base_date + timedelta(days=i * 3 + 7)).isoformat(),
            })
        
        return ExcelGenerator.create_with_dependencies(
            tasks=all_tasks,
            dependencies=[],
            program_id=program_id
        )
    
    @staticmethod
    def create_invalid_excel(error_type: str = "missing_columns") -> bytes:
        """
        Create invalid Excel for testing error handling.
        
        Args:
            error_type: Type of error to simulate:
                - "missing_columns": Missing required columns
                - "invalid_dates": Invalid date format
                - "empty_file": No data rows
                - "wrong_format": Not an Excel file
                
        Returns:
            Excel file as bytes (or other content for wrong_format)
        """
        if error_type == "wrong_format":
            return b"This is not an Excel file"
        
        wb = Workbook()
        ws = wb.active
        
        if error_type == "missing_columns":
            # Missing required Work Item ID column
            ws.cell(row=1, column=1, value="Program ID")
            ws.cell(row=1, column=2, value="Task Name")
            ws.cell(row=2, column=1, value="PROG-001")
            ws.cell(row=2, column=2, value="Test Task")
        
        elif error_type == "invalid_dates":
            headers = ["Program ID", "Work Item ID", "Work Item Name", "Planned Start", "Planned End"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            
            ws.cell(row=2, column=1, value="PROG-001")
            ws.cell(row=2, column=2, value="TASK-001")
            ws.cell(row=2, column=3, value="Test Task")
            ws.cell(row=2, column=4, value="not-a-date")
            ws.cell(row=2, column=5, value="also-not-a-date")
        
        elif error_type == "empty_file":
            headers = ["Program ID", "Work Item ID", "Work Item Name", "Planned Start", "Planned End"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            # No data rows
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
